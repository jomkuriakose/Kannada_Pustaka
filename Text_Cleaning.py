#!/usr/bin/env python
# coding: utf-8

############################################################
#Author : Bhagyashree, Jom
#Date : 1st Sept, 2020
#Purpose : Text Cleaning
#Input : Text file after timestamp removal
#Output : Text file after cleaning data
############################################################

import nltk
import numpy
import xlrd
import openpyxl 
import re
import sys
import statistics

input_file = sys.argv[1]
mapping_file = sys.argv[2]
output_file = sys.argv[3]

#input_file = 'kannada_10th_economics_ebook.txt'
#mapping_file = 'kannada_word_mapping.xlsx'
#output_file = 'tts_synth_text_output.txt'

print('\nReading input file...')
file1 = open(input_file,"r+",encoding='utf-8') 
data = file1.read()
file1.close()

print('Reading mapping file...')
wb_obj = openpyxl.load_workbook(mapping_file) 
sheet_obj = wb_obj.active 

#replace_text='\n'
replace_text=' <split> '

print('Cleaning text...')
data = data.replace('?',replace_text)
data = data.replace('_',replace_text)
data = data.replace(';',replace_text)
data = data.replace(')',replace_text)
data = data.replace('(',replace_text)
data = data.replace('!',replace_text)
data = data.replace(' – ',replace_text)
data = data.replace('-',replace_text)
data = data.replace('।',replace_text)
data = data.replace('&',replace_text)
data = data.replace('“',replace_text)
data = data.replace('”',replace_text)
data = data.replace('’',replace_text)
data = data.replace('‘',replace_text)
data = data.replace('"',replace_text)
data = data.replace(':',replace_text)
data = data.replace(',',replace_text)
data = data.replace('/',replace_text)
data = data.replace(',',replace_text)
data = data.replace('.',replace_text)
data = data.replace(' ',replace_text)
data = data.replace('|',replace_text)
data = data.replace('\t',replace_text)
data = re.sub(' +',' ',data)
data = re.sub('\n+','\n',data)
data = re.sub('\n +\n','\n',data)

m_row = sheet_obj.max_row
line = []
for i in data.split("\n"):
		line.append(i.strip())
line = '\n'.join(line)

print('Replacing using mapping file...\n')
# Replace using kannada mapping file
for i in range(1,m_row+1):
		num = sheet_obj.cell(row = i, column = 1).value
		word = sheet_obj.cell(row = i, column = 2).value
		line = line.replace(str(num), word)

# Print non-kannada words
print('\nList of word outside the Kannada Unicode range:')
uniq_words = set(data.replace("\n"," ").split(" "))
for i in uniq_words:
		for j in range(len(i)):
				if ord(i[j]) not in range(ord(u'\u0c80'), ord(u'\u0cff')+1):
						print(i)
						break
print('\n(1) Verify if the word is valid, if not correct it in the text and run again. Also check for possible multiple occurances of the word')
print('(2) If the word is valid, add the correct Kannada mapping of the word to the mapping file and run again.\n')

# Merge short sentences
# Not sure about the number

# Statistics
print('\nStatistics')
line_len = []
for i in line.split("\n"):
		line_len.append(len(i.split(" ")))
print('Longest line length (in word numbers): ',max(line_len))
print('Number of time longest line occurs: ',line_len.count(max(line_len)))
print('Shortest line length (in word numbers): ',min(line_len))
print('Number of time shortest line occurs: ',line_len.count(min(line_len)))
print('Mean of line lengths (in word numbers): ',statistics.mean(line_len))
print('SD of line lengths (in word numbers): ',statistics.stdev(line_len),'\n')

print('Total number of lines: ',len(line_len))
step = 10
line_range = [*range(min(line_len),max(line_len),step)]
line_range.append(max(line_len)+1)
for i in range(len(line_range)-1):
		print('Number of lines between lengths (',line_range[i],'-',line_range[i+1]-1,'): ',len(list(x for x in line_len if line_range[i] <= x <= line_range[i+1]-1)))

# Write to output file
print('\nWriting to output file...')
file1 = open(output_file,"w+",encoding='utf-8') 
file1.write(line)
file1.close()
print('Done!!!\n')

