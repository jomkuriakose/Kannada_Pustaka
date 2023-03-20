#!/usr/bin/python3

# Kannada TTS text cleaning
# Author: Jom Kuriakose
# email: jom@cse.iitm.ac.in
# Date: 16/02/2021

import sys
import nltk
import numpy
import xlrd
import openpyxl 
import re
import statistics

input_file = sys.argv[1]
mapping_file = sys.argv[2]
output_file = sys.argv[3]

# Comment
#input_file = 'kannada_10th_economics_ebook.txt'
#mapping_file = 'kannada_word_mapping.xlsx'
#output_file = 'tts_synth_text_output.txt'

print('\nReading input file...')
file1 = open(input_file,"r+",encoding='utf-8') 
data = file1.read()
file1.close()
line = []
# Strip each line in the data
for i in data.split("\n"):
		line.append(i.strip())
line = '\n'.join(line)

print('Reading mapping file...')
wb_obj = openpyxl.load_workbook(mapping_file) 
sheet_obj = wb_obj.active 
m_row = sheet_obj.max_row

print('Replacing numbers using mapping file...\n')
# Replace using kannada mapping file
for i in range(1,m_row+1):
		num = sheet_obj.cell(row = i, column = 1).value
		word = sheet_obj.cell(row = i, column = 2).value
		line = line.replace(str(num), word)

#
print('Cleaning text...')

replace_text=' '
replace_text_newline='\n'

#line = line.replace('?',replace_text_newline)
#line = line.replace('!',replace_text_newline)
#line = line.replace(',',replace_text_newline)
#line = line.replace('.',replace_text_newline)
#line = line.replace(';',replace_text_newline)
line = line.replace('_',replace_text)
line = line.replace(')',replace_text)
line = line.replace('(',replace_text)
line = line.replace(' – ',replace_text)
line = line.replace('-',replace_text)
line = line.replace('।',replace_text)
line = line.replace('&',replace_text)
line = line.replace('“',replace_text)
line = line.replace('”',replace_text)
line = line.replace('’',replace_text)
line = line.replace('‘',replace_text)
line = line.replace('"',replace_text)
line = line.replace(':',replace_text)
line = line.replace('/',replace_text)
line = line.replace(' ',replace_text)
line = line.replace('|',replace_text)
line = line.replace('\t',replace_text)
line = re.sub(' +',' ',line)
line = re.sub('\n+','\n',line)
line = re.sub('\n +\n','\n',line)

# Check the number of words in each line and merge with the closest small line
min_len = 3
# max_len = 10 # not used as of now.

#end_sym_list = ['.',',','?',';','!']

line_split = line.split('\n')

for line_idx in range(0,len(line_split)):
	#if(any(map(line_split[line_idx].__contains__, end_sym_list))):
		#temp_line = re.split('\. | , | ? | ; | !',line_split[line_idx])
	
	if("." in line_split[line_idx]):
		temp_line = re.split("\.",line_split[line_idx])
		flags = len(temp_line)*[0]
		flag_idx = []
		word_len = len(temp_line)*[0]
		num_long_segs = 0
		#temp_line.reverse()
		for i in range(0,len(temp_line)):
			if not(len(temp_line[i].split(' ')) <= min_len):
				flags[i] = 1
				flag_idx.append(i)
				num_long_segs = num_long_segs + 1
			word_len[i] = len(temp_line[i].split(' '))
		#temp_line.reverse()
		num_segs = 1
		for i in range(0,len(temp_line)):
			if (flags[i] == 1) and (num_segs < num_long_segs):
				num_segs = num_segs + 1
				temp_line[i] = temp_line[i]+'\n'
		line_split[line_idx] = " ".join(temp_line)
		#line_split[line_idx] = re.sub(' +',' ',line_split[line_idx])
		
	if("?" in line_split[line_idx]):
		temp_line = re.split("\?",line_split[line_idx])
		flags = len(temp_line)*[0]
		flag_idx = []
		word_len = len(temp_line)*[0]
		num_long_segs = 0
		#temp_line.reverse()
		for i in range(0,len(temp_line)):
			if not(len(temp_line[i].split(' ')) <= min_len):
				flags[i] = 1
				flag_idx.append(i)
				num_long_segs = num_long_segs + 1
			word_len[i] = len(temp_line[i].split(' '))
		#temp_line.reverse()
		num_segs = 1
		for i in range(0,len(temp_line)):
			if (flags[i] == 1) and (num_segs < num_long_segs):
				num_segs = num_segs + 1
				temp_line[i] = temp_line[i]+'\n'
		line_split[line_idx] = " ".join(temp_line)
		#line_split[line_idx] = re.sub(' +',' ',line_split[line_idx])
		
	if("!" in line_split[line_idx]):
		temp_line = re.split("\!",line_split[line_idx])
		flags = len(temp_line)*[0]
		flag_idx = []
		word_len = len(temp_line)*[0]
		num_long_segs = 0
		#temp_line.reverse()
		for i in range(0,len(temp_line)):
			if not(len(temp_line[i].split(' ')) <= min_len):
				flags[i] = 1
				flag_idx.append(i)
				num_long_segs = num_long_segs + 1
			word_len[i] = len(temp_line[i].split(' '))
		#temp_line.reverse()
		num_segs = 1
		for i in range(0,len(temp_line)):
			if (flags[i] == 1) and (num_segs < num_long_segs):
				num_segs = num_segs + 1
				temp_line[i] = temp_line[i]+'\n'
		line_split[line_idx] = " ".join(temp_line)
		#line_split[line_idx] = re.sub(' +',' ',line_split[line_idx])
	
	if("," in line_split[line_idx]):
		temp_line = re.split(",",line_split[line_idx])
		flags = len(temp_line)*[0]
		flag_idx = []
		word_len = len(temp_line)*[0]
		num_long_segs = 0
		#temp_line.reverse()
		for i in range(0,len(temp_line)):
			if not(len(temp_line[i].split(' ')) <= min_len):
				flags[i] = 1
				flag_idx.append(i)
				num_long_segs = num_long_segs + 1
			word_len[i] = len(temp_line[i].split(' '))
		#temp_line.reverse()
		num_segs = 1
		for i in range(0,len(temp_line)):
			if (flags[i] == 1) and (num_segs < num_long_segs):
				num_segs = num_segs + 1
				temp_line[i] = temp_line[i]+'\n'
		line_split[line_idx] = " ".join(temp_line)
		#line_split[line_idx] = re.sub(' +',' ',line_split[line_idx])
	
	if(";" in line_split[line_idx]):
		temp_line = re.split(";",line_split[line_idx])
		flags = len(temp_line)*[0]
		flag_idx = []
		word_len = len(temp_line)*[0]
		num_long_segs = 0
		#temp_line.reverse()
		for i in range(0,len(temp_line)):
			if not(len(temp_line[i].split(' ')) <= min_len):
				flags[i] = 1
				flag_idx.append(i)
				num_long_segs = num_long_segs + 1
			word_len[i] = len(temp_line[i].split(' '))
		#temp_line.reverse()
		num_segs = 1
		for i in range(0,len(temp_line)):
			if (flags[i] == 1) and (num_segs < num_long_segs):
				num_segs = num_segs + 1
				temp_line[i] = temp_line[i]+'\n'
		line_split[line_idx] = " ".join(temp_line)
		#line_split[line_idx] = re.sub(' +',' ',line_split[line_idx])

line = '\n'.join(line_split)
line = re.sub('\n+','\n',line)
line = re.sub(' +',' ',line)
line = re.sub(' \n ','\n',line)

# Print non-kannada words
print('\nList of word outside the Kannada Unicode range:')
uniq_words = set(line.replace("\n"," ").split(" "))
for i in uniq_words:
		for j in range(len(i)):
				if ord(i[j]) not in range(ord(u'\u0c80'), ord(u'\u0cff')+1):
						print(i)
						break
print('\n(1) Verify if the word is valid, if not correct it in the text and run again. Also check for possible multiple occurances of the word')
print('(2) If the word is valid, add the correct Kannada mapping of the word to the mapping file and run again.\n')

# Statistics
print('\nStatistics')
line_len = []
for i in line.split("\n"):
		line_len.append(len(i.split(" ")))
print('Longest line length (in word numbers): ',max(line_len))
print('Number of time longest line occurs: ',line_len.count(max(line_len)))
print('Shortest line length (in word numbers): ',min(line_len))
print('Number of time shortest line occurs: ',line_len.count(min(line_len)))
print('Mean of line lengths (in word numbers): ',statistics.mean(line_len))
print('SD of line lengths (in word numbers): ',statistics.stdev(line_len),'\n')

print('Total number of lines: ',len(line_len))
step = 10
line_range = [*range(min(line_len),max(line_len),step)]
line_range.append(max(line_len)+1)
for i in range(len(line_range)-1):
		print('Number of lines between lengths (',line_range[i],'-',line_range[i+1]-1,'): ',len(list(x for x in line_len if line_range[i] <= x <= line_range[i+1]-1)))

# Write to output file
print('\nWriting to output file...')
file1 = open(output_file,"w+",encoding='utf-8') 
file1.write(line)
file1.close()
print('Done!!!\n')

